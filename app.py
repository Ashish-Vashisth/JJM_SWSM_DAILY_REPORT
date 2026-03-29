
import re
from io import BytesIO, StringIO
from datetime import datetime

import pandas as pd
import streamlit as st

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# ===========================
# Branding: Background + Top-left watermark (Base64, JPG)
# ===========================
# Paste your Base64 (JPG) strings below (inside quotes)
BACKGROUND_B64 = """/9j/4AAQSkZ...."""

# ---------------------------
# Branding / UI CSS (IMPROVED VISIBILITY)
# ---------------------------
def apply_branding(
    bg_overlay_opacity: float = 0.28,   # lower = MORE background visibility
    card_opacity: float = 0.42          # lower = card less dark (still readable)
):
    st.markdown(
        f"""
        <style>
        /* =========================
           APP BACKGROUND (IMAGE)
           ========================= */
        [data-testid="stAppViewContainer"] {{
            background-image:
                linear-gradient(rgba(0,0,0,{bg_overlay_opacity}), rgba(0,0,0,{bg_overlay_opacity})),
                url("data:image/jpeg;base64,{BACKGROUND_B64}");
            background-size: cover;
            background-position: center;
            background-repeat: no-repeat;
            background-attachment: fixed;
        }}

        /* =========================
           MAIN CONTENT CARD
           ========================= */
        [data-testid="stAppViewContainer"] .block-container {{
            padding-top: 2.2rem;
            padding-bottom: 2rem;
            background: rgba(0, 0, 0, {card_opacity});
            border-radius: 14px;
            backdrop-filter: blur(7px);
            -webkit-backdrop-filter: blur(7px);
            border: 1px solid rgba(255,255,255,0.18);
            box-shadow: 0 12px 28px rgba(0,0,0,0.40);
            overflow: visible !important;
        }}

        /* =========================
           TEXT
           ========================= */
        h1, h2, h3, h4, h5, h6, p, label,
        .stMarkdown, .stText, .stTitle, .stSubheader, .stCaption {{
            color: #F5F6F7 !important;
        }}

        /* =========================
           INPUTS
           ========================= */
        input, textarea, select {{
            background-color: rgba(255,255,255,0.96) !important;
            color: #111 !important;
            border-radius: 10px !important;
        }}
        [data-testid="stNumberInput"] input {{
            background-color: rgba(255,255,255,0.96) !important;
            color: #111 !important;
        }}

        /* =========================
           FILE UPLOADER AREA
           ========================= */
        [data-testid="stFileUploader"] section {{
            background: rgba(255,255,255,0.18) !important;
            border-radius: 12px !important;
            border: 1px solid rgba(255,255,255,0.30) !important;
            padding: 12px !important;
            padding-bottom: 18px !important;
        }}
        [data-testid="stFileUploader"] section * {{
            color: #F8FAFC !important;
            font-weight: 600 !important;
        }}

        /* Browse files button */
        [data-testid="stFileUploader"] section button {{
            background: #ffffff !important;
            color: #111111 !important;
            font-weight: 800 !important;
            border-radius: 10px !important;
            border: 2px solid rgba(0,0,0,0.10) !important;
            box-shadow: 0 8px 18px rgba(0,0,0,0.30) !important;
            padding: 0.45rem 1.00rem !important;
        }}
        [data-testid="stFileUploader"] section button:hover {{
            background: #f1f5f9 !important;
        }}
        [data-testid="stFileUploader"] section button * {{
            color: #111111 !important;
        }}

        /* =========================
           UPLOADED FILE STRIP (LIGHTER + CLEARER)
           ========================= */
        [data-testid="stFileUploaderFile"],
        [data-testid="stFileUploader"] li {{
            background: rgba(255,255,255,0.16) !important;
            border: 1px solid rgba(255,255,255,0.22) !important;
            border-radius: 14px !important;
            padding: 12px 14px !important;
            box-shadow: 0 10px 22px rgba(0,0,0,0.25) !important;
            backdrop-filter: blur(6px);
            -webkit-backdrop-filter: blur(6px);
        }}

        /* Filename text inside strip */
        [data-testid="stFileUploaderFile"] span,
        [data-testid="stFileUploaderFile"] p,
        [data-testid="stFileUploaderFile"] small {{
            color: #F8FAFC !important;
            font-weight: 700 !important;
        }}

        /* =========================
           STRIP CLOSE (X) BUTTON — CLEAN WHITE ROUND BUTTON
           ========================= */
        [data-testid="stFileUploaderFile"] button {{
            background: #ffffff !important;
            border-radius: 12px !important;
            border: 1px solid rgba(0,0,0,0.10) !important;
            box-shadow: 0 8px 18px rgba(0,0,0,0.20) !important;
            width: 44px !important;
            height: 44px !important;
            padding: 0 !important;
        }}
        [data-testid="stFileUploaderFile"] button:hover {{
            background: #f1f5f9 !important;
        }}
        /* Make the X icon dark */
        [data-testid="stFileUploaderFile"] button svg path,
        [data-testid="stFileUploaderFile"] button svg line,
        [data-testid="stFileUploaderFile"] button svg polyline {{
            stroke: #111111 !important;
            stroke-width: 2 !important;
        }}

        /* =========================
           ALERTS
           ========================= */
        [data-testid="stAlert"] {{
            border-radius: 12px !important;
            padding: 0.75rem 1rem !important;
            margin: 0.6rem 0 !important;
            box-shadow: 0 8px 18px rgba(0,0,0,0.35) !important;
            border: 1px solid rgba(255,255,255,0.12) !important;
        }}

        /* =========================
           GENERATE REPORT BUTTON (RED)
           ========================= */
        .stButton > button {{
            background: #ff4b4b !important;
            color: #ffffff !important;
            font-weight: 900 !important;
            font-size: 1.06rem !important;
            border-radius: 14px !important;
            padding: 0.80rem 1.45rem !important;
            border: none !important;
            box-shadow: 0 10px 22px rgba(0,0,0,0.30) !important;
        }}
        .stButton > button:hover {{
            background: #e63d3d !important;
        }}

        /* =========================
           DOWNLOAD BUTTON — WHITE (TEXT + ICON ALWAYS VISIBLE)
           ========================= */
        .stDownloadButton > button {{
            background: #ffffff !important;
            color: #111111 !important;
            font-weight: 900 !important;
            border-radius: 14px !important;
            border: 2px solid rgba(0,0,0,0.10) !important;
            box-shadow: 0 12px 26px rgba(0,0,0,0.28) !important;
            padding: 0.85rem 1.35rem !important;
        }}
        .stDownloadButton > button * {{
            color: #111111 !important;
        }}
        .stDownloadButton > button svg path,
        .stDownloadButton > button svg line,
        .stDownloadButton > button svg polyline {{
            stroke: #111111 !important;
        }}

        .stDownloadButton > button:hover {{
            background: #f8fafc !important;
        }}

        /* Disabled state (still readable) */
        .stDownloadButton > button:disabled {{
            opacity: 1 !important;
            background: rgba(255,255,255,0.85) !important;
            color: rgba(17,17,17,0.85) !important;
            border: 2px solid rgba(0,0,0,0.06) !important;
            box-shadow: none !important;
            cursor: not-allowed !important;
        }}
        .stDownloadButton > button:disabled * {{
            color: rgba(17,17,17,0.70) !important;
        }}
        .stDownloadButton > button:disabled svg path,
        .stDownloadButton > button:disabled svg line,
        .stDownloadButton > button:disabled svg polyline {{
            stroke: rgba(17,17,17,0.70) !important;
        }}

        /* =========================
           EXPANDERS READABILITY
           ========================= */
        details {{
            background: rgba(15, 23, 42, 0.45) !important;
            border-radius: 12px !important;
            border: 1px solid rgba(255,255,255,0.14) !important;
            padding: 0.15rem 0.25rem !important;
        }}
        details summary {{
            background: rgba(15, 23, 42, 0.62) !important;
            border-radius: 10px !important;
            padding: 0.6rem 0.8rem !important;
            border: 1px solid rgba(255,255,255,0.18) !important;
            color: #f8fafc !important;
        }}
        </style>
        """,
        unsafe_allow_html=True
    )


# ---------------------------
# Reading the uploaded file
# ---------------------------
def read_source(uploaded_file) -> pd.DataFrame:
    """
    Robust reader:
    1) Try Excel via openpyxl (works for .xlsx/.xlsm)
    2) If fails, try HTML-table fallback (common for some .xls exports that are HTML)
    """
    raw = uploaded_file.getvalue()

    try:
        return pd.read_excel(BytesIO(raw), engine="openpyxl")
    except Exception:
        pass

    html = raw.decode("utf-8", errors="ignore")
    tables = pd.read_html(StringIO(html))
    if not tables:
        raise ValueError("Could not parse any tables from the uploaded file.")
    df = max(tables, key=lambda t: t.shape[0])
    return df


def flatten_columns(df: pd.DataFrame) -> pd.DataFrame:
    if isinstance(df.columns, pd.MultiIndex):
        df.columns = [
            " ".join([str(x) for x in tup if pd.notna(x)]).strip()
            for tup in df.columns
        ]
    return df


def normalize_columns(df: pd.DataFrame) -> dict:
    return {c: re.sub(r"\s+", "", str(c)).strip().lower() for c in df.columns}


def find_col_contains(norm_map: dict, *needles: str) -> str:
    for c, cn in norm_map.items():
        if all(n in cn for n in needles):
            return c
    raise KeyError(f"Missing column with fragments: {needles}")


# ---------------------------
# Business logic (UPDATED)
# ---------------------------
def build_report(df: pd.DataFrame, threshold: float = 75.0):
    """
    Returns: less_df, zero_df, today_zero_df

    - less_df: supplied water < threshold (based on Yesterday / Demand)
    - zero_df: ZERO(INACTIVE SITES) = Yesterday==0 AND Today==0 (excluding both blank)
    - today_zero_df: TODAY ZERO SITES = Today==0 OR blank/NaN (regardless of yesterday)
    """
    df = flatten_columns(df)
    norm = normalize_columns(df)

    scheme_id_col = find_col_contains(norm, "schemeid")
    scheme_name_col = find_col_contains(norm, "schemename")
    daily_demand_col = find_col_contains(norm, "waterdemand", "meter3", "daily")

    yest_prod_col = None
    for c, cn in norm.items():
        if ("oht" in cn) and ("watersupply" in cn) and ("meter3" in cn) and ("yesterday" in cn):
            yest_prod_col = c
            break
    if yest_prod_col is None:
        raise KeyError("Could not find 'OHT Water Supply (Meter3) Yesterday' column.")

    today_prod_col = find_col_contains(norm, "today", "waterproduction", "meter3")
    last_date_col = find_col_contains(norm, "lastdatareceivedate")

    work_df = df[[scheme_id_col, scheme_name_col, daily_demand_col, yest_prod_col, today_prod_col]].copy()
    work_df.columns = [
        "Scheme Id",
        "Scheme Name",
        "Daily Water Demand (m^3)",
        "Yesterday Water Production (m^3)",
        "Today Water Production (m^3)",
    ]

    for c in ["Daily Water Demand (m^3)", "Yesterday Water Production (m^3)", "Today Water Production (m^3)"]:
        work_df[c] = pd.to_numeric(work_df[c], errors="coerce")

    work_df["Percentage"] = (work_df["Yesterday Water Production (m^3)"] / work_df["Daily Water Demand (m^3)"]) * 100

    less_df = work_df[work_df["Percentage"].fillna(0) < threshold].copy()
    less_df["Supplied Water Percentage"] = f"<{threshold:g}%"
    less_df.insert(0, "SR.No.", range(1, len(less_df) + 1))
    less_df = less_df.drop(columns=["Today Water Production (m^3)"])

    valid_scheme = (
        work_df["Scheme Id"].notna()
        & work_df["Scheme Name"].notna()
        & (work_df["Scheme Id"].astype(str).str.strip().str.lower() != "none")
        & (work_df["Scheme Name"].astype(str).str.strip().str.lower() != "none")
        & (work_df["Scheme Id"].astype(str).str.strip() != "")
        & (work_df["Scheme Name"].astype(str).str.strip() != "")
    )

    de_blank = work_df["Yesterday Water Production (m^3)"].isna() & work_df["Today Water Production (m^3)"].isna()

    zero_mask = (
        valid_scheme
        & (~de_blank)
        & (work_df["Yesterday Water Production (m^3)"].fillna(0) == 0)
        & (work_df["Today Water Production (m^3)"].fillna(0) == 0)
    )

    zero_df = work_df.loc[zero_mask, [
        "Scheme Id",
        "Scheme Name",
        "Yesterday Water Production (m^3)",
        "Today Water Production (m^3)"
    ]].copy()

    zero_df["Last Data Receive Date"] = df.loc[zero_df.index, last_date_col].values
    zero_df["Site Status"] = "ZERO/INACTIVE SITE"
    zero_df.insert(0, "SR.No.", range(1, len(zero_df) + 1))

    today_zero_mask = valid_scheme & (work_df["Today Water Production (m^3)"].fillna(0) == 0)

    today_zero_df = work_df.loc[today_zero_mask, [
        "Scheme Id",
        "Scheme Name",
        "Today Water Production (m^3)"
    ]].copy()

    today_zero_df["Last Data Receive Date"] = df.loc[today_zero_df.index, last_date_col].values
    today_zero_df["Site Status"] = "ZERO/INACTIVE SITE"
    today_zero_df.insert(0, "SR.No.", range(1, len(today_zero_df) + 1))

    today_zero_df = today_zero_df[[
        "SR.No.",
        "Scheme Id",
        "Scheme Name",
        "Today Water Production (m^3)",
        "Last Data Receive Date",
        "Site Status",
    ]]

    return less_df, zero_df, today_zero_df


def build_lpcd_status(df: pd.DataFrame) -> pd.DataFrame:
    df = flatten_columns(df)

    if df.shape[1] < 20:
        raise ValueError(
            f"Source file has only {df.shape[1]} columns. Need at least 20 columns to extract A,B,C,R,S,T."
        )

    lpcd_df = df.iloc[:, [0, 1, 2, 17, 18, 19]].copy()
    lpcd_df.columns = [
        "Sno.",
        "Scheme Id",
        "Scheme Name",
        "Avg LPCD (Yesterday)",
        "Avg LPCD (Weekly)",
        "Avg LPCD (Monthly)",
    ]

    for c in ["Avg LPCD (Yesterday)", "Avg LPCD (Weekly)", "Avg LPCD (Monthly)"]:
        lpcd_df[c] = pd.to_numeric(lpcd_df[c], errors="coerce")

    return lpcd_df


def build_abnormal_sites(df: pd.DataFrame) -> pd.DataFrame:
    """
    Creates ABNORMAL SITES sheet with only those sites having at least one abnormal value.
    Blank/NaN source values are kept blank (not treated as abnormal display values).
    """
    df = flatten_columns(df)
    norm = normalize_columns(df)

    sno_col = df.columns[0]
    scheme_id_col = find_col_contains(norm, "schemeid")
    scheme_name_col = find_col_contains(norm, "schemename")

    hydro_col = find_col_contains(norm, "groundwaterdepth", "avg", "meter")

    radar_col = None
    for c, cn in norm.items():
        if "ohtlevel" in cn and "valueinm" in cn:
            radar_col = c
            break
    if radar_col is None:
        raise KeyError("Could not find 'OHT Level (Value in M)' column.")

    pressure_col = find_col_contains(norm, "pressure", "bar")
    turbidity_col = find_col_contains(norm, "turbidity", "ntu")
    voltage_col = find_col_contains(norm, "voltagern")

    abnormal_df = df[[
        sno_col,
        scheme_id_col,
        scheme_name_col,
        hydro_col,
        radar_col,
        pressure_col,
        turbidity_col,
        voltage_col,
    ]].copy()

    abnormal_df.columns = [
        "Sr.no",
        "Scheme Id",
        "Scheme Name",
        "Abnormal Hydrostatic Level",
        "Abnormal Radar Level",
        "Abnormal Pressure(BAR) Reading",
        "Abnormal Turbidity (NTU)",
        "Abnormal Voltage",
    ]

    abnormal_cols = [
        "Abnormal Hydrostatic Level",
        "Abnormal Radar Level",
        "Abnormal Pressure(BAR) Reading",
        "Abnormal Turbidity (NTU)",
        "Abnormal Voltage",
    ]

    for c in abnormal_cols:
        abnormal_df[c] = pd.to_numeric(abnormal_df[c], errors="coerce")

    hydro_vals = abnormal_df["Abnormal Hydrostatic Level"]
    radar_vals = abnormal_df["Abnormal Radar Level"]
    pressure_vals = abnormal_df["Abnormal Pressure(BAR) Reading"]
    turbidity_vals = abnormal_df["Abnormal Turbidity (NTU)"]
    voltage_vals = abnormal_df["Abnormal Voltage"]

    hydro_abnormal = hydro_vals.notna() & ~hydro_vals.between(18, 22.5, inclusive="both")
    radar_abnormal = radar_vals.notna() & ~((radar_vals > 0) & (radar_vals <= 4.5))
    pressure_abnormal = pressure_vals.notna() & ~pressure_vals.between(1.45, 1.95, inclusive="both")
    turbidity_abnormal = turbidity_vals.notna() & ~((turbidity_vals > 0) & (turbidity_vals <= 5))
    voltage_abnormal = voltage_vals.notna() & ((voltage_vals <= 0) | (voltage_vals < 215) | (voltage_vals > 225))

    abnormal_df.loc[~hydro_abnormal, "Abnormal Hydrostatic Level"] = pd.NA
    abnormal_df.loc[~radar_abnormal, "Abnormal Radar Level"] = pd.NA
    abnormal_df.loc[~pressure_abnormal, "Abnormal Pressure(BAR) Reading"] = pd.NA
    abnormal_df.loc[~turbidity_abnormal, "Abnormal Turbidity (NTU)"] = pd.NA
    abnormal_df.loc[~voltage_abnormal, "Abnormal Voltage"] = pd.NA

    at_least_one_abnormal = abnormal_df[abnormal_cols].notna().any(axis=1)
    abnormal_df = abnormal_df.loc[at_least_one_abnormal].copy()
    abnormal_df.reset_index(drop=True, inplace=True)

    return abnormal_df


# ---------------------------
# Excel writing + formatting
# ---------------------------
def apply_formatting(xlsx_bytes: bytes) -> bytes:
    wb = load_workbook(BytesIO(xlsx_bytes))

    thin = Side(style="thin", color="000000")
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

    align_center = Alignment(horizontal="center", vertical="center", wrap_text=False)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=False)

    header_font = Font(bold=True, color="000000")
    header_fill = PatternFill("solid", fgColor="5B9BD5")

    abnormal_fill = PatternFill("solid", fgColor="FFC7CE")
    note_label_fill = PatternFill("solid", fgColor="D9EAF7")
    note_value_fill = PatternFill("solid", fgColor="FFF2CC")
    note_font = Font(bold=True, color="000000")

    def format_sheet(ws):
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = align_center
            cell.border = border_all

        maxlen = {}
        for r in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in r:
                cell.border = border_all
                cell.alignment = align_left if cell.column == 3 else align_center
                val = "" if cell.value is None else str(cell.value)
                maxlen[cell.column] = max(maxlen.get(cell.column, 0), len(val))

        for c in range(1, ws.max_column + 1):
            header_val = str(ws.cell(row=1, column=c).value or "")
            maxlen[c] = max(maxlen.get(c, 0), len(header_val))
            width = maxlen.get(c, 0)
            ws.column_dimensions[get_column_letter(c)].width = max(10, min(60, int(width * 1.2) + 2))

    for sheet in ["LPCD STATUS", "SUPPLIED WATER LESS THAN 75", "ZERO(INACTIVE SITES)", "TODAY ZERO SITES"]:
        if sheet in wb.sheetnames:
            format_sheet(wb[sheet])

    if "ABNORMAL SITES" in wb.sheetnames:
        ws = wb["ABNORMAL SITES"]
        format_sheet(ws)

        # Highlight abnormal values in columns D:H (only nonblank abnormal entries)
        for row in range(2, ws.max_row + 1):
            for col in range(4, 9):
                cell = ws.cell(row=row, column=col)
                if cell.value not in (None, ""):
                    cell.fill = abnormal_fill
                    cell.font = note_font

        # Add normal/acceptable ranges at the bottom (excluding voltage)
        start_row = ws.max_row + 2
        notes = [
            ("Normal Hydrostatic Level", "18 to 22.5"),
            ("Normal Radar Level", "0+ to 4.5"),
            ("Normal Pressure(BAR) Reading", "1.45 to 1.95"),
            ("Normal Turbidity(NTU)", "0+ to 5"),
        ]

        for i, (label, value) in enumerate(notes, start=0):
            r = start_row + i
            label_cell = ws.cell(row=r, column=1, value=label)
            value_cell = ws.cell(row=r, column=2, value=value)

            label_cell.font = note_font
            value_cell.font = note_font
            label_cell.fill = note_label_fill
            value_cell.fill = note_value_fill

            label_cell.alignment = align_left
            value_cell.alignment = align_center

            label_cell.border = border_all
            value_cell.border = border_all

        ws.column_dimensions["A"].width = max(ws.column_dimensions["A"].width or 10, 32)
        ws.column_dimensions["B"].width = max(ws.column_dimensions["B"].width or 10, 18)

    out = BytesIO()
    wb.save(out)
    return out.getvalue()


def create_output_excel(
    less_df: pd.DataFrame,
    zero_df: pd.DataFrame,
    today_zero_df: pd.DataFrame,
    lpcd_df: pd.DataFrame,
    abnormal_df: pd.DataFrame,
) -> tuple[str, bytes]:
    date_str = datetime.now().strftime("%Y-%m-%d")
    out_name = f"ZERO & SUPPLY LESS THAN THRESHOLD SITES {date_str}.xlsx"

    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as w:
        lpcd_df.to_excel(w, sheet_name="LPCD STATUS", index=False)
        less_df.to_excel(w, sheet_name="SUPPLIED WATER LESS THAN 75", index=False)
        zero_df.to_excel(w, sheet_name="ZERO(INACTIVE SITES)", index=False)
        today_zero_df.to_excel(w, sheet_name="TODAY ZERO SITES", index=False)
        abnormal_df.to_excel(w, sheet_name="ABNORMAL SITES", index=False)

    styled = apply_formatting(buffer.getvalue())
    return out_name, styled


# ---------------------------
# Streamlit UI
# ---------------------------
st.set_page_config(page_title="UMPESL JJM SWSM Daily Report", layout="wide")
apply_branding()

st.title("UMPESL JJM SWSM Daily Report Generator")
st.write("Upload JJMUP export (.xls/.xlsx) → Download the formatted report Excel.")

threshold = st.number_input(
    "Threshold (%) for SITES LESS THAN list",
    min_value=1.0,
    max_value=100.0,
    value=75.0,
    step=1.0
)

uploaded = st.file_uploader("Upload JJMUP file", type=["xls", "xlsx", "xlsm"])

if uploaded is not None:
    st.info(f"Uploaded: {uploaded.name}")

    if st.button("Generate Report", type="primary"):
        try:
            df = read_source(uploaded)

            less_df, zero_df, today_zero_df = build_report(df, threshold=threshold)
            lpcd_df = build_lpcd_status(df)
            abnormal_df = build_abnormal_sites(df)

            out_name, out_bytes = create_output_excel(
                less_df, zero_df, today_zero_df, lpcd_df, abnormal_df
            )

            st.success(f"Created: {out_name}")

            c1, c2, c3, c4 = st.columns(4)
            c1.metric(f"SITES < {threshold:g}%", len(less_df))
            c2.metric("ZERO/INACTIVE SITES", len(zero_df))
            c3.metric("TODAY ZERO SITES", len(today_zero_df))
            c4.metric("ABNORMAL SITES", len(abnormal_df))

            with st.expander("Preview: LPCD STATUS"):
                st.dataframe(lpcd_df, use_container_width=True)

            with st.expander("Preview: SUPPLIED WATER LESS THAN THRESHOLD"):
                st.dataframe(less_df, use_container_width=True)

            with st.expander("Preview: ZERO(INACTIVE SITES)"):
                st.dataframe(zero_df, use_container_width=True)

            with st.expander("Preview: TODAY ZERO SITES"):
                st.dataframe(today_zero_df, use_container_width=True)

            with st.expander("Preview: ABNORMAL SITES"):
                st.dataframe(abnormal_df, use_container_width=True)

            st.download_button(
                "⬇️ Download Excel Report",
                data=out_bytes,
                file_name=out_name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

        except Exception as e:
            st.error("Error while generating report. Please check the uploaded file format/columns.")
            st.exception(e)

else:
    st.warning("Please upload the JJMUP export file to proceed.")

import ast
ast.parse(modified_code)
print('SYNTAX_OK')
print(len(modified_code))
